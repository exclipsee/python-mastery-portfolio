from __future__ import annotations

from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typer.testing import CliRunner

from python_mastery_portfolio.cli import app
from python_mastery_portfolio.excel_tools import write_rows_to_excel


def test_write_rows_to_excel(tmp_path: Path) -> None:
    rows = [["Name", "Score"], ["Alice", "90"], ["Bob", "88"]]
    target = tmp_path / "report.xlsx"
    out = write_rows_to_excel(rows, target)
    assert out.exists()
    wb = load_workbook(out)
    ws = cast(Worksheet, wb.active)
    assert ws["A1"].value == "Name"
    assert ws["B2"].value == "90"


def test_cli_excel_export(tmp_path: Path) -> None:
    runner = CliRunner()
    output = tmp_path / "out.xlsx"
    result = runner.invoke(
        app,
        [
            "excel-export",
            "--output",
            str(output),
            "Name,Score",
            "Alice,90",
            "Bob,88",
        ],
    )
    assert result.exit_code == 0
    assert Path(result.output.strip()).exists()


def test_write_rows_to_excel_empty(tmp_path: Path) -> None:
    target = tmp_path / "empty.xlsx"
    out = write_rows_to_excel([], target)
    assert out.exists()
