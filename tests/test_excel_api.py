from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from python_mastery_portfolio.api import app


def test_excel_export_endpoint() -> None:
    client = TestClient(app)
    rows = [["Name", "Score"], ["Alice", "90"], ["Bob", "88"]]
    r = client.post("/excel/export", json={"rows": rows})
    assert r.status_code == 200
    assert (
        r.headers.get("content-type")
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'attachment; filename="export.xlsx"' in r.headers.get("content-disposition", "")
    # Load workbook from bytes and verify contents
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws is not None, "Active worksheet should not be None"
    assert ws["A1"].value == "Name"
    assert ws["B2"].value == "90"


def test_excel_export_empty_rows_returns_400() -> None:
    client = TestClient(app)
    r = client.post("/excel/export", json={"rows": []})
    assert r.status_code == 400


def test_excel_export_non_list_row_returns_400() -> None:
    client = TestClient(app)
    # header as a string rather than a list
    r = client.post("/excel/export", json={"rows": "not-a-list"})
    assert r.status_code == 422 or r.status_code == 400


def test_excel_export_invalid_cell_type_returns_400() -> None:
    client = TestClient(app)
    # include an invalid cell type (e.g., dict)
    rows = [["Name", "Score"], ["Alice", {"bad": "cell"}]]
    r = client.post("/excel/export", json={"rows": rows})
    assert r.status_code == 400

